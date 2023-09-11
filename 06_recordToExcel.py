import os
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Side

def check_or_create_excel_file(file_path):
    if not os.path.exists(file_path):
        # Create a new workbook
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        print("New file created: {}".format(file_path))
    else:
        print("File already exists: {}".format(file_path))

def check_and_add_header(file_path):
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        if sheet.cell(row=1, column=1).value is None:
            header_row = ['No.', 'IMG_Original', 'Filename', 'IMG_DetectNum', 'Output', 'Confidence', 'IMG_DetectDial', 'Output', 'Timestamp']
            sheet.insert_rows(1)
            for col, header in enumerate(header_row, start=1):
                sheet.cell(row=1, column=col).value = header
                column_letter = openpyxl.utils.get_column_letter(col)
                column_width = len(header) + 2  # Adjust the additional width if needed
                sheet.column_dimensions[column_letter].width = column_width
            workbook.save(file_path)
            print("Header row added.")
        else:
            print("Header row already exists.")
    else:
        print("File not found: {}".format(file_path))

def add_data(file_path, img_original, filename, img_detectnumber, output_number, confidence, img_detectdial, output_dial, timestamp):
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        next_row = sheet.max_row + 1
        auto_run_number = next_row - 1
        sheet.cell(row=next_row, column=1).value = auto_run_number

        # Insert image into cells
        img_orig = Image(img_original)
        img_orig.width = 160
        img_orig.height = 180
        sheet.column_dimensions['B'].width = 200
        sheet.row_dimensions[next_row].height = 135
        img_orig.anchor = 'B{}'.format(next_row)  # Set the anchor point of the image
        sheet.add_image(img_orig)

        img_number = Image(img_detectnumber)
        img_number.width = 160
        img_number.height = 180
        sheet.column_dimensions['D'].width = 200
        sheet.row_dimensions[next_row].height = 135
        img_number.anchor = 'D{}'.format(next_row)  # Set the anchor point of the image
        sheet.add_image(img_number)

        img_dial = Image(img_detectdial)
        img_dial.width = 160
        img_dial.height = 180
        sheet.column_dimensions['G'].width = 200
        sheet.row_dimensions[next_row].height = 135
        img_dial.anchor = 'G{}'.format(next_row)  # Set the anchor point of the image
        sheet.add_image(img_dial)

        sheet.cell(row=next_row, column=3).value = filename
        sheet.cell(row=next_row, column=5).value = output_number
        sheet.cell(row=next_row, column=6).value = confidence
        sheet.cell(row=next_row, column=8).value = output_dial
        sheet.cell(row=next_row, column=9).value = timestamp

        # AutoFit column width for text cells
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 5.5) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Center align text in all cells
        for row in sheet.iter_rows(min_row=1, max_row=next_row, min_col=1, max_col=9):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Add borders to cells
                border = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
                cell.border = border

        workbook.save(file_path)
        print("Data added.")
    else:
        print("File not found: {}".format(file_path))

# Example usage
date = datetime.now().strftime("%d-%m-%Y")
folder_path = r"C:\CiRA-CORE\[detect_watermeter_cira]\logs\docs"
filename = "{}.xlsx".format(date)
file_path = os.path.join(folder_path, filename)

check_or_create_excel_file(file_path)
check_and_add_header(file_path)

# Add data to the next row
filename = payload['file_run_number']
output_number = payload['number_digi']
confidence = payload['number_confi']
output_dial = payload['dial']
timestamp = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

img_original = r"C:\CiRA-CORE\[detect_watermeter_cira]\logs\image\original_image\{}.jpg".format(filename)
img_detectnumber = r"C:\CiRA-CORE\[detect_watermeter_cira]\logs\image\detect_number\{}.jpg".format(filename)
img_detectdial = r"C:\CiRA-CORE\[detect_watermeter_cira]\logs\image\detect_dial\{}.jpg".format(filename)

add_data(file_path, img_original, filename, img_detectnumber, output_number, confidence, img_detectdial, output_dial, timestamp)

print("New file created: {}".format(file_path))
